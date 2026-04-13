from io import BytesIO
from datetime import datetime
from uuid import UUID
from typing import List, Optional

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import and_

from app.models.section_model import SectionDB
from app.models.session_model import SessionDB
from app.models.attendance_model import AttendanceRecordDB
from app.models.enrollment_model import EnrollmentDB
from app.models.student_profile_model import StudentProfileDB
from app.models.user_model import UserDB
from app.schemas.attendance_schema import AttendanceExportData, AttendanceExportRow


def format_datetime(value: Optional[datetime]) -> str:
    if not value:
        return ""
    return value.strftime("%B %d, %Y %H:%M:%S")


def auto_fit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_index = column_cells[0].column
        column_letter = get_column_letter(column_index)

        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2


def build_attendance_export_data(
    session_id: UUID,
    db: Session,
) -> AttendanceExportData:
    session = (
        db.query(SessionDB)
        .filter(
            SessionDB.SessionID == session_id,
            SessionDB.isDeleted == False,
        )
        .first()
    )
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    section = (
        db.query(SectionDB)
        .filter(
            SectionDB.SectionID == session.SectionID,
            SectionDB.isDeleted == False,
        )
        .first()
    )
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")

    enrollment_rows = (
        db.query(UserDB, AttendanceRecordDB)
        .join(
            StudentProfileDB,
            and_(
                StudentProfileDB.userID == UserDB.userID,
                StudentProfileDB.isDeleted == False,
            ),
        )
        .join(
            EnrollmentDB,
            and_(
                EnrollmentDB.StudentID == StudentProfileDB.userID,
                EnrollmentDB.SectionID == section.SectionID,
                EnrollmentDB.isDeleted == False,
            ),
        )
        .outerjoin(
            AttendanceRecordDB,
            and_(
                AttendanceRecordDB.studentUserID == StudentProfileDB.userID,
                AttendanceRecordDB.SessionID == session_id,
                AttendanceRecordDB.isDeleted == False,
            ),
        )
        .order_by(UserDB.fullName.asc())
        .all()
    )

    rows: List[AttendanceExportRow] = []
    for student, attendance in enrollment_rows:
        rows.append(
            AttendanceExportRow(
                student_user_id=student.userID,
                student_code=student.username,
                student_name=student.fullName,
                student_email=student.email,
                check_in_time=attendance.CreateAt if attendance else None,
                status=attendance.status if attendance else "Absent",
            )
        )

    attendance_count = sum(1 for row in rows if row.status in ["Present", "Late"])

    return AttendanceExportData(
        section_code=section.code,
        section_name=section.name,
        session_name=session.Name,
        session_time=session.Time,
        total_students=len(rows),
        attendance_count=attendance_count,
        rows=rows,
    )


def generate_attendance_excel(export_data: AttendanceExportData) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # ===== Styles =====
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)

    title_fill = PatternFill("solid", fgColor="0F8A9D")
    header_fill = PatternFill("solid", fgColor="0F8A9D")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # ===== Header info =====
    ws["A1"] = f"{export_data.session_name} Attendance"
    ws["A1"].font = title_font

    ws["A2"] = export_data.section_name
    ws["A3"] = f"Section Code: {export_data.section_code}"
    ws["A4"] = f"Date: {format_datetime(export_data.session_time)}"

    present_count = sum(1 for row in export_data.rows if row.status == "Present")
    absent_count = sum(1 for row in export_data.rows if row.status == "Absent")
    late_count = sum(1 for row in export_data.rows if row.status == "Late")

    ws["A6"] = f"Present: {present_count}"
    ws["B6"] = f"Absent: {absent_count}"
    ws["C6"] = f"Late: {late_count}"

    ws["A1"].fill = title_fill
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")

    # merge title row
    ws.merge_cells("A1:F1")

    # ===== Table header =====
    start_row = 8
    headers = ["#", "Student Code", "Student Name", "Email", "Time Check-in", "Status"]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # ===== Table data =====
    for index, row in enumerate(export_data.rows, start=1):
        current_row = start_row + index

        ws.cell(row=current_row, column=1, value=index).alignment = center_align
        ws.cell(row=current_row, column=2, value=row.student_code or "").alignment = left_align
        ws.cell(row=current_row, column=3, value=row.student_name).alignment = left_align
        ws.cell(row=current_row, column=4, value=row.student_email or "").alignment = left_align
        ws.cell(
            row=current_row,
            column=5,
            value=format_datetime(row.check_in_time),
        ).alignment = left_align
        ws.cell(row=current_row, column=6, value=row.status).alignment = center_align

    # ===== Bold some metadata =====
    for cell_ref in ["A2", "A3", "A4", "A6", "B6", "C6"]:
        ws[cell_ref].font = bold_font

    # ===== Freeze header =====
    ws.freeze_panes = "A9"

    # ===== Auto width =====
    auto_fit_columns(ws)

    # ===== Save to memory =====
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_attendance_service(
    session_id: UUID,
    db: Session,
) -> tuple[BytesIO, str]:
    export_data = build_attendance_export_data(session_id=session_id, db=db)
    excel_file = generate_attendance_excel(export_data)

    safe_section_code = export_data.section_code.replace(" ", "_")
    safe_session_name = export_data.session_name.replace(" ", "_")
    filename = f"attendance_{safe_section_code}_{safe_session_name}.xlsx"

    return excel_file, filename