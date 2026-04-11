import io
from datetime import datetime, timezone
from uuid import UUID

from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from app.models.attendance_model import AttendanceRecordDB
from app.models.enrollment_model import EnrollmentDB
from app.models.qrcode_model import QRCodeDB
from app.models.session_model import SessionDB
from app.schemas.attendance_schema import (
    AttendanceRecordCheckInRequest,
    AttendanceRecordUpdateRequest,
)


def check_in(
    payload: AttendanceRecordCheckInRequest,
    student_id: UUID,
    db: Session,
) -> AttendanceRecordDB:
    # 1. Tìm QR token hợp lệ (còn active, chưa hết hạn)
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    qr = (
        db.query(QRCodeDB)
        .filter(
            QRCodeDB.token == payload.token,
            QRCodeDB.isActive == True,
            QRCodeDB.expireAt > now,
        )
        .first()
    )
    if not qr:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid or expired QR code.",
        )

    # 2. Kiểm tra session tồn tại và chưa bị xoá
    session = (
        db.query(SessionDB)
        .filter(SessionDB.SessionID == qr.SessionID, SessionDB.isDeleted == False)
        .first()
    )
    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Session not found.",
        )

    # 3. Kiểm tra student đã enrolled vào section của session này chưa
    enrolled = (
        db.query(EnrollmentDB)
        .filter(
            EnrollmentDB.SectionID == session.SectionID,
            EnrollmentDB.StudentID == student_id,
            EnrollmentDB.isDeleted == False,
        )
        .first()
    )
    if not enrolled:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You are not enrolled in this section.",
        )

    # 4. Kiểm tra đã điểm danh cho session này chưa
    existing = (
        db.query(AttendanceRecordDB)
        .filter(
            AttendanceRecordDB.SessionID == session.SessionID,
            AttendanceRecordDB.studentUserID == student_id,
            AttendanceRecordDB.isDeleted == False,
        )
        .first()
    )
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="You have already checked in for this session.",
        )

    # 5. Tạo attendance record
    record = AttendanceRecordDB(
        studentUserID=student_id,
        SessionID=session.SessionID,
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    return record


def list_session_attendance(
    session_id: UUID,
    db: Session,
) -> list[AttendanceRecordDB]:
    """Return all attendance records for a session, with student info eager-loaded."""
    # Verify session exists
    session = (
        db.query(SessionDB)
        .filter(SessionDB.SessionID == session_id, SessionDB.isDeleted == False)
        .first()
    )
    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Session not found.",
        )

    records = (
        db.query(AttendanceRecordDB)
        .options(joinedload(AttendanceRecordDB.student).joinedload("user"))
        .filter(
            AttendanceRecordDB.SessionID == session_id,
            AttendanceRecordDB.isDeleted == False,
        )
        .order_by(AttendanceRecordDB.CreateAt)
        .all()
    )
    return records


def export_session_attendance(
    session_id: UUID,
    db: Session,
) -> StreamingResponse:
    """Export attendance records for a session as an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl is not installed. Run: pip install openpyxl",
        )

    # Verify session exists
    session = (
        db.query(SessionDB)
        .filter(SessionDB.SessionID == session_id, SessionDB.isDeleted == False)
        .first()
    )
    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Session not found.",
        )

    records = (
        db.query(AttendanceRecordDB)
        .options(joinedload(AttendanceRecordDB.student).joinedload("user"))
        .filter(
            AttendanceRecordDB.SessionID == session_id,
            AttendanceRecordDB.isDeleted == False,
        )
        .order_by(AttendanceRecordDB.CreateAt)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center = Alignment(horizontal="center")

    headers = ["#", "Student ID", "Full Name", "Username", "Email", "Check-in Time"]
    ws.append(headers)
    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Data rows
    for idx, record in enumerate(records, start=1):
        student_user = None
        if record.student and hasattr(record.student, "user"):
            student_user = record.student.user

        ws.append([
            idx,
            str(record.studentUserID),
            student_user.fullName if student_user else "",
            student_user.username if student_user else "",
            student_user.email if student_user else "",
            record.CreateAt.strftime("%Y-%m-%d %H:%M:%S") if record.CreateAt else "",
        ])

    # Auto-size columns
    for column in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in column), default=0)
        ws.column_dimensions[column[0].column_letter].width = max(max_length + 4, 12)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"attendance_session_{session_id}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def update_attendance_record(
    attendance_record_id: UUID,
    payload: AttendanceRecordUpdateRequest,
    db: Session,
) -> AttendanceRecordDB:
    """Soft-delete or restore an attendance record (teacher use)."""
    record = (
        db.query(AttendanceRecordDB)
        .filter(AttendanceRecordDB.AttendanceRecordID == attendance_record_id)
        .first()
    )
    if not record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance record not found.",
        )

    record.isDeleted = payload.isDeleted
    db.commit()
    db.refresh(record)
    return record